from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
from datetime import datetime
import time

def obter_hora_processamento():
    hora_atual = ""
    hora_atual = datetime.fromtimestamp(time.time()).strftime('%H:%M:%S:%M')
    return(hora_atual)

def gerar_excel(data_excel):

    # definir o caminho da pasta
    caminho = "C:/temp/nfe_excel/"
    nome_arquivo = 'relatorio_nfe.xlsx'
    caminho_completo = caminho + nome_arquivo
    total = 0
    # Cria um novo arquivo do Excel
    workbook = Workbook()

    # Seleciona a primeira planilha do arquivo
    worksheet = workbook.active

    # Escreve o cabeçalho na primeira linha do Excel
    worksheet.cell(row=1, column=1, value='Data')
    worksheet.cell(row=1, column=2, value='Produto')
    worksheet.cell(row=1, column=3, value='Valor')

    # Escreve os dados no Excel
    linha = 2
    for item in data_excel:
        coluna = 1
        for valor in item:
            worksheet.cell(row=linha, column=coluna, value=valor)
            coluna += 1
        linha += 1

    for item in data_excel:
        valor1 = float(item[2])
        #valor1 = item[2].replace("R$ ", "")
        #print(valor1)
        #valor1 = float(valor1.replace(",", ""))
        #print(valor1)
        total += valor1
        worksheet.cell(row=linha, column=2, value="Total")
        worksheet.cell(row=linha, column=3, value=total)

    print("A soma do campo de valor é: R$", total)
    # Salva o arquivo do Excel
    workbook.save(filename=caminho_completo)
    print("Excel gerado com sucesso!")

def format_xml_portalfiscal(data_excel):
    # definir o caminho da pasta
    # http://www.portalfiscal.inf.br/nfe

    caminho = "C:/temp/xml_nfe_portal/"
    caminho_completo = ""
    total_registros = 0

    # obter uma lista de todos os arquivos na pasta
    lista_arquivos = os.listdir(caminho) 

    print("Quantidade de Notas DANFE..: ", len(lista_arquivos))

    for arquivo in lista_arquivos:
        caminho_completo = caminho + arquivo
        print("Nota Lida...: ", arquivo)
        with open(caminho_completo, 'r') as f:
            data = f.read()
        bs_data = BeautifulSoup(data, 'xml')
        b_unique = bs_data.find_all('det')

        ide = bs_data.find('ide')
        dhEmi = ide.find('dhEmi').text
        dt = datetime.strptime(dhEmi, "%Y-%m-%dT%H:%M:%S%z")
        data_formatada = dt.strftime("%d/%m/%Y")

        for item in b_unique:
            produto = item.find('xProd').text
            valor = float(item.find('vProd').text)
            #valor_formatado = "R$ {:.2f}".format(valor)
            data_registro = (data_formatada, produto, valor)
            data_excel.append(data_registro)
            total_registros = total_registros + 1

    return  total_registros           

def format_xml_nfe_barueri(data_excel):

    # http://www.barueri.sp.gov.br/nfe

    caminho = "C:/temp/xml_nfe_barueri/"
    caminho_completo = ""
    total_registros = 0
    data_formatada = ""

    # obter uma lista de todos os arquivos na pasta
    lista_arquivos = os.listdir(caminho) 

    print("Quantidade de Nota Barueri..: ", len(lista_arquivos))

    for arquivo in lista_arquivos:
        caminho_completo = caminho + arquivo
        print("Nota Lida...: ", arquivo)

        with open(caminho_completo, 'r') as f:
            data = f.read()
        bs_data = BeautifulSoup(data, 'xml')
        
        ide = bs_data.find('InfNfeServPrestado')
        dhEmi = ide.find('DataEmissao').text
        dt = datetime.strptime(dhEmi, "%Y-%m-%dT%H:%M:%S")
        data_formatada = dt.strftime("%d/%m/%Y")

        ide = bs_data.find('ServicoPrestado')
        produto = ide.find('DescricaoServico').text

        ide = bs_data.find('ValoresNfe')
        valor = float(ide.find('ValorLiquidoNfe').text)
        #valor_formatado = "R$ {:.2f}".format(valor)

        data_registro = (data_formatada, produto, valor)
        data_excel.append(data_registro)

        total_registros = total_registros + 1

    return total_registros

if __name__ == '__main__':
    # código para ser executado quando o módulo é executado 
    # como programa principal
    data_excel = []
    total_registros = 0

    obter_hora_processamento()
    print("Inicio do Processamento...: ", obter_hora_processamento())
    qdt_portalfiscal = format_xml_portalfiscal(data_excel)
    qdt_nfe_barueri = format_xml_nfe_barueri(data_excel)
    total_registros = qdt_portalfiscal + qdt_nfe_barueri
    print("Total de Itens a ser gravado...: ", total_registros)
    gerar_excel(data_excel)
    print("Fim do Processamento...: ", obter_hora_processamento())
